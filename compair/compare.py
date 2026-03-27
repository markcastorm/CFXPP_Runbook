"""
CFXPP Comparison Tool
=====================
Powerful standalone script to compare pipeline output against reference files.

Usage:
    1. Edit config.py to set paths
    2. Run: python compare.py

Generates:
    - comparison_report.csv: Detailed row-by-row comparison
    - comparison_summary.txt: Executive summary with statistics
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import csv
import os
import sys
import shutil
from datetime import datetime
from collections import defaultdict, Counter

import config


class ComparisonReport:
    """Handles comparison between pipeline output and reference data."""

    def __init__(self):
        self.output_codes = []
        self.output_cells = {}
        self.reference_codes = []
        self.reference_cells = {}
        self.code_to_entries = {}
        self.descriptions = {}

        self.matches = []
        self.mismatches = []
        self.missing = []  # Reference has, pipeline doesn't
        self.extra = []    # Pipeline has, reference doesn't

        # Track cell coordinates for highlighting
        self.mismatch_coords = []  # (date, col_idx) tuples
        self.missing_coords = []
        self.extra_coords = []

    def load_excel_data(self, path, label):
        """Load column codes and cell values from DATA xlsx."""
        if config.VERBOSE:
            print(f'Loading {label}: {os.path.basename(path)}')

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['DATA']

        # Load column codes (row 1, cols 2-541)
        codes = []
        for c in range(2, 542):
            code = ws.cell(row=1, column=c).value
            codes.append(code)

        # Load cell values
        cells = {}
        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(row=r, column=1).value
            if date_val is None:
                continue

            # Format date
            if hasattr(date_val, 'strftime'):
                d = date_val.strftime('%Y-%m-%d')
            else:
                d = str(date_val)

            # Load all data columns
            for c in range(2, 542):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    cells[(d, c - 2)] = v

        wb.close()

        if config.VERBOSE:
            print(f'  Loaded: {len(codes)} codes, {len(cells)} cells')

        return codes, cells

    def load_mapping_csv(self, path):
        """Load source-column mapping CSV."""
        if config.VERBOSE:
            print(f'Loading source mapping: {os.path.basename(path)}')

        code_to_entries = defaultdict(list)

        if not os.path.exists(path):
            print(f'  WARNING: Mapping CSV not found: {path}')
            return code_to_entries

        with open(path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                code_to_entries[row['Output_Column_Code']].append(row)

        if config.VERBOSE:
            print(f'  Loaded: {len(code_to_entries)} column mappings')

        return code_to_entries

    def load_meta_descriptions(self, output_data_path):
        """Load column descriptions from META file."""
        meta_path = output_data_path.replace('DATA', 'META')
        descs = {}

        if not os.path.exists(meta_path):
            if config.VERBOSE:
                print(f'  No META file found: {os.path.basename(meta_path)}')
            return descs

        if config.VERBOSE:
            print(f'Loading META descriptions: {os.path.basename(meta_path)}')

        wb = openpyxl.load_workbook(meta_path, data_only=True)
        ws = wb.active

        for r in range(2, ws.max_row + 1):
            code = ws.cell(row=r, column=1).value
            desc = ws.cell(row=r, column=2).value
            if code:
                descs[code] = desc or ''

        wb.close()

        if config.VERBOSE:
            print(f'  Loaded: {len(descs)} descriptions')

        return descs

    def get_section(self, code):
        """Determine which section a column belongs to."""
        if not code:
            return 'UNKNOWN'

        if '.CURRENCYPOSITIONING.' in code and '.G10.' in code:
            return 'G10 Currency Positioning'
        elif '.CURRENCYPOSITIONING.' in code and '.EM.' in code:
            return 'EM Currency Positioning'
        elif '.FXPAIRPOSITIONING.' in code:
            return 'FX Pair Positioning'

        return 'OTHER'

    def parse_column_metadata(self, code):
        """Parse pair, client, and metric from a column code."""
        if not code:
            return '', '', ''

        parts = code.split('.')

        if '.FXPAIRPOSITIONING.' in code:
            # FX Pair format
            pair = parts[-2] if len(parts) >= 2 else ''

            if 'VOLUME_NORMALIZED' in code:
                metric = 'Volume (normalized)'
            elif 'CLOSING_PRICE' in code:
                metric = 'Closing Price'
            else:
                metric = ''

            # Client type
            if 'BANKS_BROKERVOLUME_NORMALIZED' in code:
                client = 'BANKS_BROKER'
            else:
                client = parts[3] if len(parts) > 3 else ''

            return pair, client, metric

        elif '.CURRENCYPOSITIONING.' in code:
            # Currency Positioning format
            ccy = parts[-2] if len(parts) >= 2 else ''
            group = parts[-3] if len(parts) >= 3 else ''
            client = parts[-4] if len(parts) >= 4 else ''

            return f'{group}/{ccy}', client, 'Net Cumulative Positioning'

        return '', '', ''

    def compare_values(self, val1, val2):
        """Compare two values with tolerance for floats."""
        if val1 is None and val2 is None:
            return True, 0

        if val1 is None or val2 is None:
            return False, None

        try:
            f1 = float(val1)
            f2 = float(val2)
            diff = f1 - f2

            if abs(diff) < config.FLOAT_TOLERANCE:
                return True, 0

            return False, diff

        except (ValueError, TypeError):
            # String comparison
            return str(val1) == str(val2), None

    def build_row_dict(self, status, date, col_idx, code, out_val, ref_val, diff):
        """Build a standardized row dictionary for the report."""
        section = self.get_section(code)
        pair, client, metric = self.parse_column_metadata(code)
        desc = self.descriptions.get(code, '')

        # Get source files
        entries = self.code_to_entries.get(code, [])
        source_files = []
        archive_links = []
        raw_clients = []

        for e in entries:
            fn = e.get('Source_File', '')
            if fn and fn not in source_files:
                source_files.append(fn)
                if config.ARCHIVE_DIR:
                    archive_links.append(os.path.join(config.ARCHIVE_DIR, fn))

            raw = e.get('Raw_Client_Types', '')
            if raw and raw not in raw_clients:
                raw_clients.append(raw)

        return {
            'Status': status,
            'Date': date,
            'Output_Column': col_idx + 2,  # +2 because col 1 is date
            'Column_Code': code or '',
            'Section': section,
            'Currency_Pair': pair,
            'Client_Type': client,
            'Metric': metric,
            'Column_Description': desc,
            'Pipeline_Value': out_val if out_val is not None else '',
            'Reference_Value': ref_val if ref_val is not None else '',
            'Difference': round(diff, 4) if diff is not None else '',
            'Source_File(s)': '; '.join(source_files) if source_files else 'NO SOURCE FILE',
            'Archive_Link(s)': '; '.join(archive_links) if archive_links else '',
            'Raw_Client_Types': '; '.join(raw_clients) if raw_clients else '',
        }

    def perform_comparison(self):
        """Main comparison logic."""
        if config.VERBOSE:
            print('\nPerforming comparison...')

        # Combine all keys from both datasets
        all_keys = set(self.output_cells.keys()) | set(self.reference_cells.keys())

        if config.VERBOSE:
            print(f'  Total cell locations: {len(all_keys)}')

        for key in sorted(all_keys):
            date, col_idx = key
            out_val = self.output_cells.get(key)
            ref_val = self.reference_cells.get(key)

            code = self.output_codes[col_idx] if col_idx < len(self.output_codes) else ''

            if out_val is None and ref_val is None:
                # Both empty - skip
                continue

            elif out_val is not None and ref_val is not None:
                # Both have values - compare
                match, diff = self.compare_values(out_val, ref_val)

                if match:
                    self.matches.append(key)
                else:
                    row = self.build_row_dict('MISMATCH', date, col_idx, code, out_val, ref_val, diff)
                    self.mismatches.append(row)
                    self.mismatch_coords.append(key)

            elif out_val is None:
                # Reference has value, pipeline doesn't
                row = self.build_row_dict('MISSING', date, col_idx, code, None, ref_val, None)
                self.missing.append(row)
                self.missing_coords.append(key)

            else:
                # Pipeline has value, reference doesn't
                row = self.build_row_dict('EXTRA', date, col_idx, code, out_val, None, None)
                self.extra.append(row)
                self.extra_coords.append(key)

        if config.VERBOSE:
            print(f'  Matches:    {len(self.matches)}')
            print(f'  Mismatches: {len(self.mismatches)}')
            print(f'  Missing:    {len(self.missing)}')
            print(f'  Extra:      {len(self.extra)}')

    def write_csv_report(self, output_path):
        """Write detailed CSV report."""
        if config.VERBOSE:
            print(f'\nWriting CSV report: {os.path.basename(output_path)}')

        header = [
            'Status', 'Date', 'Output_Column', 'Column_Code',
            'Section', 'Currency_Pair', 'Client_Type', 'Metric',
            'Column_Description',
            'Pipeline_Value', 'Reference_Value', 'Difference',
            'Source_File(s)', 'Archive_Link(s)', 'Raw_Client_Types',
        ]

        all_rows = self.mismatches + self.missing + self.extra

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=header)
            writer.writeheader()
            for row in all_rows:
                writer.writerow(row)

        if config.VERBOSE:
            print(f'  Written {len(all_rows)} rows')

    def generate_summary_stats(self):
        """Generate summary statistics."""
        stats = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'output_file': config.OUTPUT_DATA,
            'reference_file': config.REFERENCE_DATA,
            'total_cells_compared': len(self.matches) + len(self.mismatches) + len(self.missing) + len(self.extra),
            'matches': len(self.matches),
            'mismatches': len(self.mismatches),
            'missing': len(self.missing),
            'extra': len(self.extra),
            'match_rate': 0,
        }

        total = stats['total_cells_compared']
        if total > 0:
            stats['match_rate'] = (len(self.matches) / total) * 100

        return stats

    def analyze_by_category(self, rows, key_func):
        """Group rows by a category function."""
        groups = defaultdict(list)
        for row in rows:
            category = key_func(row)
            groups[category].append(row)
        return groups

    def write_summary_report(self, output_path):
        """Write human-readable summary report."""
        if config.VERBOSE:
            print(f'\nWriting summary report: {os.path.basename(output_path)}')

        stats = self.generate_summary_stats()

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('=' * 70 + '\n')
            f.write('CFXPP COMPARISON REPORT\n')
            f.write('=' * 70 + '\n\n')

            f.write(f"Generated: {stats['timestamp']}\n\n")

            f.write('FILES COMPARED:\n')
            f.write(f"  Pipeline Output: {os.path.basename(stats['output_file'])}\n")
            f.write(f"  Reference:       {os.path.basename(stats['reference_file'])}\n\n")

            f.write('OVERALL STATISTICS:\n')
            f.write(f"  Total cells compared: {stats['total_cells_compared']:,}\n")
            f.write(f"  Matches:              {stats['matches']:,} ({stats['match_rate']:.2f}%)\n")
            f.write(f"  Mismatches:           {stats['mismatches']:,}\n")
            f.write(f"  Missing (in ref):     {stats['missing']:,}\n")
            f.write(f"  Extra (in pipeline):  {stats['extra']:,}\n\n")

            # Mismatches by section
            if self.mismatches:
                f.write('MISMATCHES BY SECTION:\n')
                by_section = self.analyze_by_category(self.mismatches, lambda r: r['Section'])
                for section, items in sorted(by_section.items()):
                    f.write(f"  {section}: {len(items)}\n")
                f.write('\n')

                # Top mismatches by magnitude
                f.write('TOP MISMATCHES (by absolute difference):\n')
                sorted_mm = sorted(
                    [r for r in self.mismatches if r['Difference'] != ''],
                    key=lambda r: abs(float(r['Difference'])) if r['Difference'] else 0,
                    reverse=True
                )[:config.MAX_CONSOLE_EXAMPLES]

                for r in sorted_mm:
                    f.write(f"  {r['Date']} | {r['Currency_Pair']} {r['Client_Type']} {r['Metric']}\n")
                    f.write(f"    Pipeline: {r['Pipeline_Value']} | Reference: {r['Reference_Value']} | Diff: {r['Difference']}\n")
                    f.write(f"    Source: {r['Source_File(s)']}\n")
                f.write('\n')

            # Missing by pair
            if self.missing:
                f.write('MISSING CELLS BY CURRENCY PAIR:\n')
                by_pair = self.analyze_by_category(self.missing, lambda r: r['Currency_Pair'])
                for pair, items in sorted(by_pair.items(), key=lambda x: -len(x[1])):
                    has_source = any(r['Source_File(s)'] != 'NO SOURCE FILE' for r in items)
                    src_note = '' if has_source else ' [NO SOURCE FILES]'
                    f.write(f"  {pair}: {len(items)} cells{src_note}\n")
                f.write('\n')

            # Extra by pair
            if self.extra:
                f.write('EXTRA CELLS BY CURRENCY PAIR:\n')
                by_pair = self.analyze_by_category(self.extra, lambda r: r['Currency_Pair'])
                for pair, items in sorted(by_pair.items(), key=lambda x: -len(x[1])):
                    f.write(f"  {pair}: {len(items)} cells\n")
                f.write('\n')

            f.write('=' * 70 + '\n')
            f.write(f"Detailed CSV report: {os.path.basename(config.REPORT_OUTPUT or 'comparison_report.csv')}\n")
            f.write('=' * 70 + '\n')

        if config.VERBOSE:
            print(f'  Summary written')

    def create_annotated_excel(self, source_path, output_path, coords_dict, label):
        """
        Create annotated Excel file with highlighted cells.

        Args:
            source_path: Original Excel file path
            output_path: Where to save annotated version
            coords_dict: Dict mapping status -> list of (date, col_idx) tuples
            label: 'Pipeline' or 'Reference'
        """
        if config.VERBOSE:
            print(f'\nCreating annotated {label} file: {os.path.basename(output_path)}')

        # Copy and open workbook
        shutil.copy2(source_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb['DATA']

        # Define styles
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        bold_font = Font(bold=True)

        # Build date-to-row mapping
        date_to_row = {}
        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(row=r, column=1).value
            if date_val:
                if hasattr(date_val, 'strftime'):
                    d = date_val.strftime('%Y-%m-%d')
                else:
                    d = str(date_val)
                date_to_row[d] = r

        # Highlight cells
        highlight_counts = {'MISMATCH': 0, 'MISSING': 0, 'EXTRA': 0}

        for status, coords in coords_dict.items():
            if status == 'MISMATCH':
                fill = red_fill
            elif status == 'MISSING':
                fill = yellow_fill
            else:  # EXTRA
                fill = green_fill

            for date, col_idx in coords:
                row_num = date_to_row.get(date)
                if row_num:
                    col_num = col_idx + 2  # +2 because col 1 is date, data starts at col 2
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = fill
                    highlight_counts[status] += 1

        # Add legend at top
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value='LEGEND:')
        ws.cell(row=1, column=1).font = bold_font

        ws.cell(row=1, column=2, value='RED = Mismatch')
        ws.cell(row=1, column=2).fill = red_fill

        ws.cell(row=1, column=3, value='YELLOW = Missing')
        ws.cell(row=1, column=3).fill = yellow_fill

        ws.cell(row=1, column=4, value='GREEN = Extra')
        ws.cell(row=1, column=4).fill = green_fill

        ws.cell(row=1, column=5, value=f'Annotated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')

        wb.save(output_path)
        wb.close()

        if config.VERBOSE:
            print(f'  Highlighted: {highlight_counts["MISMATCH"]} mismatches (red), '
                  f'{highlight_counts["MISSING"]} missing (yellow), '
                  f'{highlight_counts["EXTRA"]} extra (green)')

    def run(self):
        """Execute full comparison workflow."""
        print('=' * 70)
        print('CFXPP COMPARISON TOOL')
        print('=' * 70)
        print()

        # Validate paths
        if not os.path.exists(config.OUTPUT_DATA):
            print(f'ERROR: Output file not found: {config.OUTPUT_DATA}')
            return 1

        if not os.path.exists(config.REFERENCE_DATA):
            print(f'ERROR: Reference file not found: {config.REFERENCE_DATA}')
            return 1

        # Load data
        self.output_codes, self.output_cells = self.load_excel_data(config.OUTPUT_DATA, 'Pipeline Output')
        self.reference_codes, self.reference_cells = self.load_excel_data(config.REFERENCE_DATA, 'Reference')

        # Verify column codes match
        if self.output_codes != self.reference_codes:
            print('\nWARNING: Column codes differ between output and reference!')
            print('This may cause incorrect comparisons.')

        # Load supporting data
        self.code_to_entries = self.load_mapping_csv(config.MAPPING_CSV)
        self.descriptions = self.load_meta_descriptions(config.OUTPUT_DATA)

        # Perform comparison
        self.perform_comparison()

        # Create timestamped output directory
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        compair_root = os.path.dirname(os.path.abspath(__file__))
        run_dir = os.path.join(compair_root, f'run_{timestamp}')
        os.makedirs(run_dir, exist_ok=True)

        if config.VERBOSE:
            print(f'\nCreating output directory: {os.path.basename(run_dir)}')

        # Determine output paths
        report_csv = os.path.join(run_dir, config.REPORT_FILENAME)
        summary_txt = os.path.join(run_dir, config.SUMMARY_FILENAME)

        # Copy original files to run directory
        if config.VERBOSE:
            print('\nCopying original files...')

        output_copy = os.path.join(run_dir, f'ORIGINAL_{os.path.basename(config.OUTPUT_DATA)}')
        ref_copy = os.path.join(run_dir, f'ORIGINAL_{os.path.basename(config.REFERENCE_DATA)}')

        shutil.copy2(config.OUTPUT_DATA, output_copy)
        shutil.copy2(config.REFERENCE_DATA, ref_copy)

        if config.VERBOSE:
            print(f'  {os.path.basename(output_copy)}')
            print(f'  {os.path.basename(ref_copy)}')

        # Create annotated versions with highlighting
        annotated_output = os.path.join(run_dir, f'ANNOTATED_Pipeline_{os.path.basename(config.OUTPUT_DATA)}')
        annotated_ref = os.path.join(run_dir, f'ANNOTATED_Reference_{os.path.basename(config.REFERENCE_DATA)}')

        # For pipeline output: highlight mismatches (red), extra (green)
        pipeline_coords = {
            'MISMATCH': self.mismatch_coords,
            'EXTRA': self.extra_coords,
        }
        self.create_annotated_excel(config.OUTPUT_DATA, annotated_output, pipeline_coords, 'Pipeline')

        # For reference: highlight mismatches (red), missing (yellow)
        reference_coords = {
            'MISMATCH': self.mismatch_coords,
            'MISSING': self.missing_coords,
        }
        self.create_annotated_excel(config.REFERENCE_DATA, annotated_ref, reference_coords, 'Reference')

        # Write reports
        self.write_csv_report(report_csv)
        self.write_summary_report(summary_txt)

        # Print summary to console
        print('\n' + '=' * 70)
        print('COMPARISON COMPLETE')
        print('=' * 70)

        stats = self.generate_summary_stats()
        print(f"\nMatches:    {stats['matches']:,} ({stats['match_rate']:.2f}%)")
        print(f"Mismatches: {stats['mismatches']:,}")
        print(f"Missing:    {stats['missing']:,}")
        print(f"Extra:      {stats['extra']:,}")

        print(f'\nOutput directory: {run_dir}')
        print(f'\nFiles created:')
        print(f"  {os.path.basename(report_csv)}")
        print(f"  {os.path.basename(summary_txt)}")
        print(f"  {os.path.basename(output_copy)}")
        print(f"  {os.path.basename(ref_copy)}")
        print(f"  {os.path.basename(annotated_output)}")
        print(f"  {os.path.basename(annotated_ref)}")
        print('=' * 70)

        return 0


def main():
    """Main entry point."""
    try:
        report = ComparisonReport()
        exit_code = report.run()
        return exit_code

    except KeyboardInterrupt:
        print('\n\nInterrupted by user')
        return 130

    except Exception as e:
        print(f'\nERROR: {e}')
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
