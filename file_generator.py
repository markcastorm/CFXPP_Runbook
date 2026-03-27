"""
CFXPP File Generator
=====================
Generates output DATA, META, and ZIP files.

This module handles:
- Creating DATA Excel files (Row 1: codes, Row 2: descriptions, Row 3+: data)
- Creating META Excel files with time series metadata
- Creating ZIP archives
- Updating the master data file
- Managing timestamped and latest output folders
"""

import os
import logging
import shutil
import zipfile

import pandas as pd
import openpyxl

import config

logger = logging.getLogger(__name__)


class CFXPPFileGenerator:
    """Generates output files for CFXPP runbook."""

    def __init__(self, column_order=None):
        """
        Args:
            column_order: List of (code, description) tuples.
        """
        self.column_order = column_order or []
        self.timestamp = config.get_timestamp()

    def _create_output_dirs(self, batch_id=None):
        """Create output directory structure."""
        ts_dir = os.path.join(config.OUTPUT_DIR, self.timestamp)
        os.makedirs(ts_dir, exist_ok=True)

        latest_dir = os.path.join(config.OUTPUT_DIR, config.LATEST_FOLDER)
        os.makedirs(latest_dir, exist_ok=True)

        os.makedirs(config.MASTER_DIR, exist_ok=True)

        return ts_dir, latest_dir

    def create_data_file(self, df, output_path):
        """
        Create DATA Excel file.

        Structure:
        - Row 1: Column codes (empty first cell, then codes)
        - Row 2: Column descriptions (empty first cell, then descriptions)
        - Row 3+: Date and data values

        Args:
            df: DataFrame with 'date' column + data columns (keyed by code).
            output_path: Path for output file.

        Returns:
            str: Path to created file.
        """
        logger.info(f'Creating DATA file: {output_path}')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'DATA'

        codes = [code for code, _ in self.column_order]
        descs = [desc for _, desc in self.column_order]

        # Row 1: codes
        for i, code in enumerate(codes, start=2):
            ws.cell(row=1, column=i, value=code)

        # Row 2: descriptions
        for i, desc in enumerate(descs, start=2):
            ws.cell(row=2, column=i, value=desc)

        # Row 3+: data
        for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
            # Date in first column
            date_val = row.get('date', '')
            ws.cell(row=row_idx, column=1, value=str(date_val))

            for col_idx, code in enumerate(codes, start=2):
                value = row.get(code)
                if value is not None and not (isinstance(value, float) and pd.isna(value)):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                # Leave cell empty for None/NaN (blank)

        wb.save(output_path)
        wb.close()

        logger.info(f'DATA file created: {len(df)} rows, {len(codes)} columns')
        return output_path

    def create_meta_file(self, output_path):
        """
        Create META Excel file with time series metadata.

        Args:
            output_path: Path for output file.

        Returns:
            str: Path to created file.
        """
        logger.info(f'Creating META file: {output_path}')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'META'

        # Write headers
        for col, header in enumerate(config.META_HEADERS, start=1):
            ws.cell(row=1, column=col, value=header)

        # Write one row per column
        for row_idx, (code, desc) in enumerate(self.column_order, start=2):
            # Determine frequency from code suffix
            if code.endswith('.B'):
                frequency = 'B'
            elif code.endswith('.D'):
                frequency = 'D'
            else:
                frequency = 'D'

            # CODE_MNEMONIC: code without trailing suffix (.B or .D)
            code_mnemonic = code
            if code_mnemonic.endswith('.B') or code_mnemonic.endswith('.D'):
                code_mnemonic = code_mnemonic[:-2]

            ws.cell(row=row_idx, column=1, value=code)
            ws.cell(row=row_idx, column=2, value=code_mnemonic)
            ws.cell(row=row_idx, column=3, value=desc)
            ws.cell(row=row_idx, column=4, value=frequency)
            ws.cell(row=row_idx, column=5, value=config.METADATA_DEFAULTS['MULTIPLIER'])
            ws.cell(row=row_idx, column=6, value=config.METADATA_DEFAULTS['AGGREGATION_TYPE'])
            ws.cell(row=row_idx, column=7, value=config.METADATA_DEFAULTS['UNIT_TYPE'])
            ws.cell(row=row_idx, column=8, value=config.METADATA_DEFAULTS['DATA_TYPE'])
            ws.cell(row=row_idx, column=9, value=config.METADATA_DEFAULTS['DATA_UNIT'])
            ws.cell(row=row_idx, column=10, value=config.METADATA_DEFAULTS['SEASONALLY_ADJUSTED'])
            ws.cell(row=row_idx, column=11, value=config.METADATA_DEFAULTS['ANNUALIZED'])
            ws.cell(row=row_idx, column=12, value=config.METADATA_DEFAULTS['PROVIDER_MEASURE_URL'])
            ws.cell(row=row_idx, column=13, value=config.METADATA_DEFAULTS['PROVIDER'])
            ws.cell(row=row_idx, column=14, value=config.METADATA_DEFAULTS['SOURCE'])
            ws.cell(row=row_idx, column=15, value=config.METADATA_DEFAULTS['SOURCE_DESCRIPTION'])
            ws.cell(row=row_idx, column=16, value=config.METADATA_DEFAULTS['COUNTRY'])
            ws.cell(row=row_idx, column=17, value=config.METADATA_DEFAULTS['DATASET'])
            ws.cell(row=row_idx, column=18, value=config.METADATA_DEFAULTS['LAST_RELEASE_DATE'])

        wb.save(output_path)
        wb.close()

        logger.info(f'META file created: {len(self.column_order)} entries')
        return output_path

    def create_zip_file(self, data_file, meta_file, zip_path):
        """
        Create ZIP archive containing DATA and META files.

        Args:
            data_file: Path to DATA file.
            meta_file: Path to META file.
            zip_path: Path for ZIP file.

        Returns:
            str: Path to created ZIP file.
        """
        logger.info(f'Creating ZIP file: {zip_path}')

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(data_file, os.path.basename(data_file))
            zf.write(meta_file, os.path.basename(meta_file))

        logger.info('ZIP file created')
        return zip_path

    def save_master_data(self, df, batch_id):
        """
        Save updated master data file for a batch.

        Args:
            df: DataFrame with all accumulated data.
            batch_id: Batch identifier string.

        Returns:
            str: Path to master file.
        """
        master_path = os.path.join(config.MASTER_DIR, f'Master_CFXPP_DATA_{batch_id}.xlsx')
        logger.info(f'Saving master data to: {master_path}')
        self.create_data_file(df, master_path)
        logger.info(f'Master data saved: {len(df)} rows')
        return master_path

    def load_master_data(self, batch_id):
        """
        Load existing master data file for a batch.

        Args:
            batch_id: Batch identifier string.

        Returns:
            pd.DataFrame or empty DataFrame if master doesn't exist.
        """
        master_path = os.path.join(config.MASTER_DIR, f'Master_CFXPP_DATA_{batch_id}.xlsx')

        if not os.path.exists(master_path):
            logger.info(f'No master data for batch {batch_id}')
            return pd.DataFrame()

        try:
            logger.info(f'Loading master data from {master_path}')
            wb = openpyxl.load_workbook(master_path, data_only=True)
            ws = wb.active

            # Read column codes from row 1
            columns = ['date']
            for col in range(2, ws.max_column + 1):
                code = ws.cell(row=1, column=col).value
                if code:
                    columns.append(str(code))
                else:
                    break

            # Read data from row 3+
            data = []
            for row in range(3, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    continue
                row_data = [str(date_val)]
                for col in range(2, len(columns) + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                data.append(row_data)

            wb.close()

            if not data:
                return pd.DataFrame(columns=columns)

            df = pd.DataFrame(data, columns=columns)
            logger.info(f'Loaded {len(df)} rows from master data')
            return df

        except Exception as e:
            logger.error(f'Error loading master data: {e}')
            return pd.DataFrame()

    def generate_files(self, df, batch_id):
        """
        Generate all output files: timestamped + latest + master.

        Args:
            df: DataFrame with combined data.
            batch_id: Batch identifier string.

        Returns:
            dict: Paths to all created files.
        """
        logger.info('Generating output files')

        ts_dir, latest_dir = self._create_output_dirs(batch_id)

        # Timestamped file names
        data_name = f'{config.DATA_FILE_PREFIX}_{self.timestamp}.xlsx'
        meta_name = f'{config.META_FILE_PREFIX}_{self.timestamp}.xlsx'
        zip_name = f'{config.ZIP_FILE_PREFIX}_{self.timestamp}.zip'

        # Latest file names
        latest_data_name = f'{config.DATA_FILE_PREFIX}_LATEST.xlsx'
        latest_meta_name = f'{config.META_FILE_PREFIX}_LATEST.xlsx'
        latest_zip_name = f'{config.ZIP_FILE_PREFIX}_LATEST.zip'

        # Full paths
        ts_data = os.path.join(ts_dir, data_name)
        ts_meta = os.path.join(ts_dir, meta_name)
        ts_zip = os.path.join(ts_dir, zip_name)

        latest_data = os.path.join(latest_dir, latest_data_name)
        latest_meta = os.path.join(latest_dir, latest_meta_name)
        latest_zip = os.path.join(latest_dir, latest_zip_name)

        # Create timestamped files
        self.create_data_file(df, ts_data)
        self.create_meta_file(ts_meta)
        self.create_zip_file(ts_data, ts_meta, ts_zip)

        # Copy to latest
        shutil.copy2(ts_data, latest_data)
        shutil.copy2(ts_meta, latest_meta)
        shutil.copy2(ts_zip, latest_zip)
        logger.info('Copied files to latest folder')

        # Update master
        master_path = self.save_master_data(df, batch_id)

        result = {
            'data_file': ts_data,
            'meta_file': ts_meta,
            'zip_file': ts_zip,
            'latest_data': latest_data,
            'latest_meta': latest_meta,
            'latest_zip': latest_zip,
            'master_file': master_path,
            'timestamp': self.timestamp,
            'output_dir': ts_dir,
        }

        logger.info('All output files generated successfully')
        return result
