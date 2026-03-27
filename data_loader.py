"""
CFXPP Data Loader
==================
Scans input directories recursively for Excel files and loads them.

This module handles:
- Recursive scanning of Input/ and all subfolders for .xlsx/.xlsm files
- Loading workbooks and reading the 'Export' sheet into raw 2D grids
- Providing a picklable load function for multiprocessing workers
"""

import os
import logging
from glob import glob

import openpyxl

import config

logger = logging.getLogger(__name__)


class CFXPPDataLoader:
    """Scans and loads Barclays FX export Excel files."""

    def __init__(self):
        self.input_dir = config.INPUT_DIR

    def find_input_files(self):
        """
        Recursively find all Excel files under the input directory.

        Returns:
            list: Sorted list of absolute file paths (newest first by mtime).
        """
        files = []
        for root, dirs, filenames in os.walk(self.input_dir):
            for fname in filenames:
                if fname.startswith('~$'):
                    continue
                if fname.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                    files.append(os.path.join(root, fname))

        files.sort(key=os.path.getmtime, reverse=True)

        logger.info(f'Found {len(files)} input file(s) under {self.input_dir}')
        return files


def load_file_raw(file_path):
    """
    Load a single Excel file and return raw sheet data as a 2D list.

    This is a module-level function (not a method) so it can be pickled
    and dispatched to ProcessPoolExecutor workers.

    Args:
        file_path: Absolute path to the Excel file.

    Returns:
        dict with:
            - 'file_path': absolute path
            - 'file_name': basename
            - 'data': 2D list of cell values (0-based row/col)
            - 'max_rows': row count
            - 'max_cols': col count
        Returns None on failure.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Find the Export sheet
        ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == 'export':
                ws = wb[sheet_name]
                break

        if ws is None:
            # Fallback: use first sheet
            ws = wb[wb.sheetnames[0]]

        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column,
                                values_only=True):
            data.append(list(row))

        max_rows = len(data)
        max_cols = len(data[0]) if data else 0

        wb.close()

        return {
            'file_path': file_path,
            'file_name': os.path.basename(file_path),
            'data': data,
            'max_rows': max_rows,
            'max_cols': max_cols,
        }

    except Exception as e:
        logger.error(f'Error loading {file_path}: {e}')
        return None


if __name__ == '__main__':
    from logger_setup import setup_logging
    setup_logging()

    loader = CFXPPDataLoader()
    files = loader.find_input_files()
    print(f'Found {len(files)} input files')

    if files:
        result = load_file_raw(files[0])
        if result:
            print(f"Loaded: {result['file_name']}")
            print(f"Size: {result['max_rows']} rows x {result['max_cols']} cols")
            print(f"\nFirst 8 rows:")
            for i, row in enumerate(result['data'][:8]):
                vals = [v for v in row if v is not None]
                print(f"  Row {i}: {vals}")
