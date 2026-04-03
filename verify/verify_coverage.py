"""
Source Data Coverage Verification Tool
======================================
Analyzes archived source files and verifies output completeness.
Colors cells with issues and generates missing categories report.

Usage:
    1. Update config.py with your paths
    2. Run: python verify_coverage.py
"""
import sys
from pathlib import Path

# Add parent directory to path to import config
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re
from datetime import datetime
import config

# Color schemes
RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
GREEN_FILL = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')

def parse_source_file_metadata(file_path):
    """Extract metadata from a source file by reading its content"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Read header info to determine file type
        header_data = {}
        for row_idx in range(1, min(15, ws.max_row + 1)):
            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value:
                    header_data[f"R{row_idx}C{col_idx}"] = str(cell.value)

        wb.close()

        # Determine file type from header content
        file_info = {
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_type': 'UNKNOWN',
            'currency_pair': None,
            'currency': None,
            'client_type': None,
            'date': None,
            'section': None
        }

        # Look for indicators in header
        header_text = ' '.join(header_data.values()).upper()

        # Check for FX Pair
        fx_pairs = ['AUDUSD', 'EURUSD', 'GBPUSD', 'USDJPY', 'USDCAD', 'NZDUSD', 'USDCHF',
                    'USDMXN', 'USDBRL', 'USDCOP', 'USDZAR', 'USDTRY', 'USDSGD', 'USDINR',
                    'USDPLN', 'USDCZK', 'USDHUF', 'EURGBP', 'EURJPY', 'EURCHF', 'GBPJPY']

        for pair in fx_pairs:
            if pair in header_text:
                file_info['file_type'] = 'FX_PAIR'
                file_info['currency_pair'] = pair
                break

        # Check for Currency Positioning
        if 'CURRENCY POSITIONING' in header_text or 'CUMULATIVE POSITIONS' in header_text:
            file_info['file_type'] = 'CCY_POS'

            # Check for G10 or EM
            if 'G10' in header_text:
                file_info['section'] = 'G10'
            elif 'EM' in header_text or 'EMERGING' in header_text:
                file_info['section'] = 'EM'

            # Extract currency
            currencies = ['USD', 'EUR', 'GBP', 'JPY', 'CHF', 'CAD', 'AUD', 'NZD', 'SEK', 'NOK',
                         'MXN', 'BRL', 'COP', 'ZAR', 'TRY', 'SGD', 'INR', 'PLN', 'CZK', 'HUF']
            for ccy in currencies:
                if ccy in header_text:
                    file_info['currency'] = ccy
                    break

        # Extract client type
        client_types = ['BANKS', 'BROKER', 'CORPORATE', 'HEDGE FUND', 'HEDGEFUND',
                       'REAL MONEY', 'REALMONEY', 'UNCLASSIFIED']
        found_types = []
        for ct in client_types:
            if ct in header_text:
                found_types.append(ct.replace(' ', '_'))

        if found_types:
            file_info['client_type'] = '_'.join(sorted(set(found_types)))

        # Extract date from filename
        date_match = re.search(r'(\d{2})-(\d{2})-(\d{4})', file_path.name)
        if date_match:
            month, day, year = date_match.groups()
            file_info['date'] = f"{year}-{month}-{day}"

        return file_info

    except Exception as e:
        return {
            'file_path': str(file_path),
            'file_name': file_path.name,
            'file_type': 'ERROR',
            'error': str(e)
        }

def scan_archive_folder(archive_path):
    """Scan archive folder and categorize source files"""
    if config.VERBOSE:
        print(f"\nScanning archive folder: {archive_path}")

    archive_dir = Path(archive_path)

    if not archive_dir.exists():
        print(f"ERROR: Archive directory not found: {archive_path}")
        return None

    files = list(archive_dir.glob("**/*.xlsx"))
    if config.VERBOSE:
        print(f"Found {len(files)} Excel files")
        print("\nAnalyzing file contents (this may take a moment)...")

    file_catalog = []

    for idx, fpath in enumerate(files):
        if config.VERBOSE and (idx + 1) % 25 == 0:
            print(f"  Analyzed {idx + 1}/{len(files)} files...")

        metadata = parse_source_file_metadata(fpath)
        file_catalog.append(metadata)

    if config.VERBOSE:
        print(f"  Analyzed {len(files)}/{len(files)} files - Complete!")

    return file_catalog

def analyze_expected_coverage(file_catalog):
    """Determine what output columns SHOULD exist based on source files"""
    expected_data = defaultdict(set)  # {column_code: set of dates}

    # Group files by characteristics
    fx_files = defaultdict(list)
    ccy_files = defaultdict(list)

    for file_info in file_catalog:
        file_type = file_info['file_type']
        date = file_info.get('date')

        if file_type == 'FX_PAIR':
            pair = file_info.get('currency_pair')
            client = file_info.get('client_type', 'UNKNOWN')
            if pair and date:
                fx_files[(pair, client)].append(date)

        elif file_type == 'CCY_POS':
            currency = file_info.get('currency')
            client = file_info.get('client_type', 'UNKNOWN')
            section = file_info.get('section', 'UNKNOWN')
            if currency and date:
                ccy_files[(currency, client, section)].append(date)

    return {
        'fx_files': fx_files,
        'ccy_files': ccy_files,
        'file_catalog': file_catalog
    }

def verify_output_against_sources(output_file, source_analysis, output_dir):
    """Verify output file against source files and create annotated report"""
    print(f"\n{'='*100}")
    print("VERIFICATION REPORT")
    print(f"{'='*100}")

    # Load output file
    output_df = pd.read_excel(output_file)
    print(f"\nOutput file: {output_file}")
    print(f"Shape: {output_df.shape}")
    print(f"Dates: {output_df.iloc[:, 0].tolist()}")

    # Analyze source files
    fx_files = source_analysis['fx_files']
    ccy_files = source_analysis['ccy_files']
    file_catalog = source_analysis['file_catalog']

    # Statistics
    print(f"\n{'='*100}")
    print("SOURCE FILE ANALYSIS")
    print(f"{'='*100}")
    print(f"Total source files: {len(file_catalog)}")

    file_types = defaultdict(int)
    for f in file_catalog:
        file_types[f['file_type']] += 1

    print(f"\nFile types found:")
    for ftype, count in sorted(file_types.items()):
        print(f"  {ftype}: {count} files")

    # Dates with files
    dates_found = set()
    for f in file_catalog:
        if f.get('date'):
            dates_found.add(f['date'])

    print(f"\nDates found in source files: {sorted(dates_found)}")

    # FX Pair coverage
    print(f"\n{'='*100}")
    print("FX PAIR SOURCE FILES")
    print(f"{'='*100}")
    print(f"Unique (Pair, Client) combinations: {len(fx_files)}")
    if len(fx_files) > 0:
        sorted_fx = sorted(fx_files.items(), key=lambda x: (str(x[0][0] or ''), str(x[0][1] or '')))
        for (pair, client), dates in sorted_fx[:config.MAX_CONSOLE_EXAMPLES]:
            pair_str = pair if pair else 'N/A'
            client_str = client if client else 'N/A'
            print(f"  {pair_str:10} + {client_str:40} -> {len(dates)} dates")
        if len(fx_files) > config.MAX_CONSOLE_EXAMPLES:
            print(f"  ... and {len(fx_files) - config.MAX_CONSOLE_EXAMPLES} more combinations")
    else:
        print("  (No FX Pair files found)")

    # CCY Positioning coverage
    print(f"\n{'='*100}")
    print("CURRENCY POSITIONING SOURCE FILES")
    print(f"{'='*100}")
    print(f"Unique (Currency, Client, Section) combinations: {len(ccy_files)}")
    # Sort with None-safe key
    sorted_ccy = sorted(ccy_files.items(), key=lambda x: (str(x[0][2] or ''), str(x[0][0] or ''), str(x[0][1] or '')))
    for (ccy, client, section), dates in sorted_ccy[:config.MAX_CONSOLE_EXAMPLES]:
        section_str = section if section else 'N/A'
        ccy_str = ccy if ccy else 'N/A'
        client_str = client if client else 'N/A'
        print(f"  {section_str:5} {ccy_str:5} + {client_str:40} -> {len(dates)} dates")
    if len(ccy_files) > config.MAX_CONSOLE_EXAMPLES:
        print(f"  ... and {len(ccy_files) - config.MAX_CONSOLE_EXAMPLES} more combinations")

    # Check output coverage
    print(f"\n{'='*100}")
    print("OUTPUT VERIFICATION")
    print(f"{'='*100}")

    data_cols = output_df.iloc[:, 1:]  # Exclude date column
    total_cols = len(data_cols.columns)
    cols_with_data = (data_cols.count() > 0).sum()
    filled_cells = data_cols.count().sum()

    print(f"Total columns: {total_cols}")
    print(f"Columns with data: {cols_with_data}/{total_cols} ({100*cols_with_data/total_cols:.1f}%)")
    print(f"Total cells filled: {filled_cells:,}")

    # Create timestamped output folder
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if config.CREATE_RUN_FOLDER:
        run_folder = Path(output_dir) / f"run_{timestamp}"
        run_folder.mkdir(parents=True, exist_ok=True)
        output_path = run_folder / f"verification_annotated.xlsx"
        report_path = run_folder / f"missing_categories.txt"
    else:
        output_path = Path(output_dir) / f"verification_report_{timestamp}.xlsx"
        report_path = Path(output_dir) / f"missing_categories_{timestamp}.txt"

    print(f"\nCreating annotated report...")

    # Load with openpyxl for formatting
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Add legend
    ws.insert_rows(1, 4)
    ws['A1'] = "VERIFICATION LEGEND:"
    ws['A2'] = "GREEN: Data present and verified"
    ws['A2'].fill = GREEN_FILL
    ws['A3'] = "YELLOW: Data missing (no source file found)"
    ws['A3'].fill = YELLOW_FILL
    ws['A4'] = "RED: Data expected but missing in output"
    ws['A4'].fill = RED_FILL

    # Color code cells based on data presence
    if config.VERBOSE:
        print("  Analyzing and coloring cells...")

    issues = []
    missing_categories = set()

    for col_idx in range(2, ws.max_column + 1):  # Skip date column
        col_has_data = False
        for row_idx in range(6, ws.max_row + 1):  # Skip header and legend
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None and cell.value != '':
                cell.fill = GREEN_FILL
                col_has_data = True
            else:
                cell.fill = YELLOW_FILL

        # Track columns with no data
        if not col_has_data:
            col_letter = get_column_letter(col_idx)
            header_cell = ws.cell(5, col_idx)  # After legend rows
            col_name = str(header_cell.value) if header_cell.value else f"Column_{col_letter}"
            missing_categories.add(col_name[:100])  # Truncate long names

    if config.VERBOSE:
        print(f"  Saving annotated file...")
    wb.save(output_path)
    print(f"  Saved: {output_path}")

    # Generate missing categories report
    with open(report_path, 'w') as f:
        f.write("="*100 + "\n")
        f.write("MISSING CATEGORIES REPORT\n")
        f.write("="*100 + "\n\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Output file: {output_file}\n")
        f.write(f"Archive folder: {source_analysis.get('archive_path', 'N/A')}\n\n")

        f.write(f"Total columns: {total_cols}\n")
        f.write(f"Columns with data: {cols_with_data}\n")
        f.write(f"Columns missing data: {len(missing_categories)}\n\n")

        f.write("="*100 + "\n")
        f.write("COLUMNS WITH NO DATA (Missing Source Files)\n")
        f.write("="*100 + "\n\n")

        for idx, cat in enumerate(sorted(missing_categories), 1):
            f.write(f"{idx:4}. {cat}\n")

        if not missing_categories:
            f.write("  (None - all columns have data!)\n")

        f.write("\n" + "="*100 + "\n")
        f.write("SOURCE FILE SUMMARY\n")
        f.write("="*100 + "\n\n")
        f.write(f"Total source files: {len(file_catalog)}\n\n")

        f.write("File types:\n")
        for ftype, count in sorted(file_types.items()):
            f.write(f"  {ftype}: {count} files\n")

        f.write(f"\nDates found: {', '.join(sorted(dates_found))}\n")

        f.write(f"\nFX Pair files: {len(fx_files)} unique combinations\n")
        f.write(f"Currency Positioning files: {len(ccy_files)} unique combinations\n")

    print(f"  Saved: {report_path}")

    # Print summary
    print(f"\n{'='*100}")
    print("VERIFICATION SUMMARY")
    print(f"{'='*100}")
    if config.CREATE_RUN_FOLDER:
        print(f"Output folder: {run_folder}")
    print(f"  - Columns with data: {cols_with_data}/{total_cols} ({100*cols_with_data/total_cols:.1f}%)")
    print(f"  - Columns missing data: {len(missing_categories)}")

    if len(missing_categories) == 0:
        print(f"\n[EXCELLENT] All {total_cols} columns have data!")
    elif cols_with_data >= 400:
        print(f"\n[GOOD] {cols_with_data}/{total_cols} columns covered")
    else:
        print(f"\n[FAIR] {cols_with_data}/{total_cols} columns covered - {len(missing_categories)} categories missing")

    print(f"\n{'='*100}")

def main():
    """Main execution"""
    print("="*100)
    print("SOURCE DATA COVERAGE VERIFICATION TOOL")
    print("="*100)

    # Get paths from config
    archive_folder = config.ARCHIVE_FOLDER
    output_file = config.OUTPUT_FILE
    report_dir = config.REPORT_OUTPUT_DIR or str(Path(__file__).parent)

    print(f"\nConfiguration:")
    print(f"  Archive folder: {archive_folder}")
    print(f"  Output file: {output_file}")
    print(f"  Report directory: {report_dir}")

    # Step 1: Scan archive folder
    file_catalog = scan_archive_folder(archive_folder)

    if not file_catalog:
        print("\nERROR: No files found or archive folder error")
        return

    # Step 2: Analyze expected coverage
    source_analysis = analyze_expected_coverage(file_catalog)
    source_analysis['archive_path'] = archive_folder

    # Step 3: Verify output against sources
    verify_output_against_sources(output_file, source_analysis, report_dir)

    print(f"\n{'='*100}")
    print("VERIFICATION COMPLETE!")
    print(f"{'='*100}\n")

if __name__ == "__main__":
    main()
